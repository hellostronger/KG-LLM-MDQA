#!/usr/bin/env python3
"""
从 test_docs.json（Excel 伪段落）构建知识图谱 excel_kg.pkl
节点类型: Sheet, MergeArea, Cell, Header
边类型  : contains / has_header / at_row / at_col / value_is
输出     : List[Dict] -> 直接可被 KG-LLM-MDQA Pipeline 使用
"""

import json, pickle, re, os
from typing import List, Dict
from collections import defaultdict
from openpyxl import load_workbook

# ---------- 1. 读取 Excel 原始文件 ----------
def load_workbook_raw(excel_path: str):
    """返回 workbook 与合并映射"""
    wb = load_workbook(excel_path, data_only=True)
    merged_map = {}
    for sheet in wb.worksheets:
        for m in sheet.merged_cells.ranges:
            r1, c1, r2, c2 = m.bounds
            val = str(sheet.cell(r1, c1).value or "")
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    merged_map[(sheet.title, r, c)] = val
    return wb, merged_map

# ---------- 2. 抽取三元 ----------
def extract_triples_from_sheet(sheet, merged_map: Dict) -> List[Dict]:
    triples = []
    sheet_id = sheet.title
    # 2.1 建立 Sheet 节点
    triples.append({
        "id": sheet_id,
        "type": "Sheet",
        "title": sheet.title,
    })

    # 2.2 扫描合并区域
    for m in sheet.merged_cells.ranges:
        r1, c1, r2, c2 = m.bounds
        area_range = f"{sheet.title}!{m.coord}"
        header_val = str(sheet.cell(r1, c1).value or "")
        # 合并区域节点
        triples.append({
            "id": area_range,
            "type": "MergeArea",
            "header": header_val,
            "min_row": r1,
            "max_row": r2,
            "min_col": c1,
            "max_col": c2,
            "sheet": sheet_id,
        })
        # 边: Sheet -> contains -> MergeArea
        triples.append({
            "src": sheet_id,
            "rel": "contains",
            "dst": area_range,
        })

        # 2.3 把区域内每个单元格都连进来
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell_val = merged_map.get((sheet.title, r, c), str(sheet.cell(r, c).value or ""))
                cell_id = f"{sheet.title}!{r}_{c}"
                triples.append({
                    "id": cell_id,
                    "type": "Cell",
                    "value": cell_val,
                    "row": r,
                    "col": c,
                })
                # 边: MergeArea -> contains -> Cell
                triples.append({
                    "src": area_range,
                    "rel": "contains",
                    "dst": cell_id,
                })

    # 2.4 把第一行当作 Header
    if sheet.max_row >= 1:
        for c in range(1, sheet.max_column + 1):
            hdr = str(sheet.cell(1, c).value or f"Col{c}")
            hdr_id = f"{sheet.title}!H{c}"
            triples.append({
                "id": hdr_id,
                "type": "Header",
                "text": hdr,
                "col": c,
            })
            # 边: Header -> describes -> Col
            for r in range(2, sheet.max_row + 1):
                cell_id = f"{sheet.title}!{r}_{c}"
                triples.append({
                    "src": hdr_id,
                    "rel": "describes",
                    "dst": cell_id,
                })
    return triples

# ---------- 3. 主流程 ----------
def build_excel_kg(docs_json: str, kg_pkl: str):
    """
    docs_json  : test_docs.json（excel2docs.py 产出）
    kg_pkl     : excel_kg.pkl（KG-LLM-MDQA Pipeline 所需）
    """
    # 3.1 收集所有 Excel 路径
    excel_files = set()
    with open(docs_json, encoding="utf-8") as f:
        docs = json.load(f)
    for d in docs:
        fid = d["id"].split("_")[0] + ".xlsx"  # 约定：id=file_sheet_...
        excel_files.add(fid)

    triples = []
    for fname in excel_files:
        fname = os.path.join("../../excel_raw", fname)
        wb, merged = load_workbook_raw(fname)
        for sheet in wb.worksheets:
            triples.extend(extract_triples_from_sheet(sheet, merged))

    # 3.2 保存
    with open(kg_pkl, "wb") as f:
        pickle.dump(triples, f)
    print(f"[build_excel_kg] 节点+边共 {len(triples)} 条 -> 已保存 {kg_pkl}")

# ---------- 4. CLI ----------
if __name__ == "__main__":
    import argparse, sys
    parser = argparse.ArgumentParser()
    parser.add_argument("--docs", default="test_docs.json")
    parser.add_argument("--out", default="excel_kg.pkl")
    args = parser.parse_args()
    if not os.path.exists(args.docs):
        sys.exit(f"{args.docs} 不存在，请先跑 excel2docs.py")
    build_excel_kg(args.docs, args.out)