import glob, json, os, re
from openpyxl import load_workbook

def excel2docs(excel_dir, out_json):
    docs=[]
    for f in glob.glob(os.path.join(excel_dir,'*.xlsx')):
        wb=load_workbook(f, data_only=True)
        file_id=os.path.basename(f).split('.')[0]
        for sheet in wb.worksheets:
            # 1. 建立合并回填映射
            merged={}
            for m in sheet.merged_cells.ranges:
                r1,c1,r2,c2=m.bounds
                v=str(sheet.cell(r1,c1).value or "")
                for r in range(r1,r2+1):
                    for c in range(c1,c2+1):
                        merged[(r,c)]=v
            # 2. 按行扫描，每 50 行拆一段落（可调）
            buf=[]
            for r,row in enumerate(sheet.iter_rows(values_only=True),1):
                line="\t".join([merged.get((r,c+1), str(v or ""))
                                for c,v in enumerate(row)])
                buf.append(line)
                if len(buf)==50:
                    docs.append({
                        "id": f"{file_id}_{sheet.title}_r{r-49}-{r}",
                        "title": f"{sheet.title} rows {r-49}-{r}",
                        "text": "\n".join(buf)
                    })
                    buf=[]
            if buf:
                docs.append({
                    "id": f"{file_id}_{sheet.title}_tail",
                    "title": f"{sheet.title} tail",
                    "text": "\n".join(buf)
                })
    json.dump(docs, open(out_json,'w',encoding='utf8'), ensure_ascii=False, indent=2)

if __name__=="__main__":
    excel2docs("../../excel_raw", "test_docs.json")